
# Importaciones necesarias para la aplicación
from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
import os
import re

# Configuración básica de Flask
app = Flask(__name__)
app.secret_key = "clave_contactos"

# Rutas y configuración de datos
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARCHIVO = os.path.join(BASE_DIR, "contactos.xlsx")
ENCABEZADOS = ["Nombre", "Apellido", "Teléfono", "Correo", "Dirección", "Categoría", "Favorito"]
CATEGORIAS = ["Familia", "Trabajo", "Amigos", "Otro"]


# Crear archivo Excel con contactos de ejemplo si no existe
def crear_excel():
    if not os.path.exists(ARCHIVO):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Contactos"
        hoja.append(ENCABEZADOS)
        hoja.append(["Ana", "Mora", "88887777", "ana@gmail.com", "San José", "Familia", "Sí"])
        hoja.append(["Luis", "Soto", "77776666", "luis@gmail.com", "Cartago", "Trabajo", "No"])
        hoja.append(["Maria", "Vargas", "66665555", "maria@gmail.com", "Heredia", "Amigos", "Sí"])
        hoja.append(["Pedro", "Rojas", "55554444", "pedro@gmail.com", "Alajuela", "Otro", "No"])
        hoja.append(["Sofia", "Castro", "44443333", "sofia@gmail.com", "Limón", "Familia", "Sí"])
        libro.save(ARCHIVO)


# Leer todos los contactos del archivo Excel
def leer_contactos():
    crear_excel()
    libro = load_workbook(ARCHIVO)
    hoja = libro.active
    contactos = []

    # Recorrer las filas del Excel y crear diccionarios con los datos
    for fila in range(2, hoja.max_row + 1):
        contactos.append({
            "id": fila,
            "nombre": hoja.cell(fila, 1).value or "",
            "apellido": hoja.cell(fila, 2).value or "",
            "telefono": str(hoja.cell(fila, 3).value or ""),
            "correo": hoja.cell(fila, 4).value or "",
            "direccion": hoja.cell(fila, 5).value or "",
            "categoria": hoja.cell(fila, 6).value or "Otro",
            "favorito": hoja.cell(fila, 7).value or "No"
        })
    return contactos


# Buscar un contacto específico por su id
def buscar_por_id(contacto_id):
    for contacto in leer_contactos():
        if contacto["id"] == contacto_id:
            return contacto
    return None


# Validar los datos del contacto
def validar(nombre, telefono, correo):
    if nombre == "" or telefono == "" or correo == "":
        return "Nombre, teléfono y correo son obligatorios."
    if len(telefono) != 8 or not telefono.isdigit():
        return "El teléfono debe tener exactamente 8 dígitos."
    if not re.match(r"^[^@]+@[^@]+\.[^@]+$", correo):
        return "El correo debe tener un formato válido."
    return ""


#Página de login
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        contrasena = request.form["contrasena"]
        # Validar credenciales (admin/1234)
        if usuario == "admin" and contrasena == "1234":
            session["usuario"] = usuario
            return redirect(url_for("contactos"))
        else:
            return render_template("login.html", error="Usuario o contraseña incorrectos")
    return render_template("login.html")


# Lista de todos los contactos
@app.route("/contactos")
def contactos():
    if "usuario" not in session:
        return redirect(url_for("login"))
    lista = leer_contactos()
    # Si se pide ordenar de A a Z
    if request.args.get("orden") == "az":
        lista.sort(key=lambda x: x["nombre"].lower())
    return render_template("contactos.html", contactos=lista)


# Agregar nuevo contacto
@app.route("/agregar", methods=["GET", "POST"])
def agregar():
    if "usuario" not in session:
        return redirect(url_for("login"))
    if request.method == "POST":
        nombre = request.form["nombre"]
        apellido = request.form["apellido"]
        telefono = request.form["telefono"]
        correo = request.form["correo"]
        direccion = request.form["direccion"]
        categoria = request.form["categoria"]
        favorito = request.form.get("favorito", "No")
        # Validar datos antes de guardar
        mensaje = validar(nombre, telefono, correo)
        if mensaje != "":
            return render_template("agregar.html", categorias=CATEGORIAS, error=mensaje)
        libro = load_workbook(ARCHIVO)
        hoja = libro.active
        hoja.append([nombre, apellido, telefono, correo, direccion, categoria, favorito])
        libro.save(ARCHIVO)
        return redirect(url_for("contactos"))
    return render_template("agregar.html", categorias=CATEGORIAS)


# Buscar contactos por nombre
@app.route("/buscar", methods=["GET", "POST"])
def buscar():
    if "usuario" not in session:
        return redirect(url_for("login"))
    resultados = []
    texto = ""
    if request.method == "POST":
        texto = request.form["texto"].lower()
        # Buscar en nombre y apellido
        for contacto in leer_contactos():
            nombre_completo = (contacto["nombre"] + " " + contacto["apellido"]).lower()
            if texto in nombre_completo:
                resultados.append(contacto)
    return render_template("buscar.html", resultados=resultados, texto=texto)


# Ver detalle de un contacto
@app.route("/detalle/<int:contacto_id>")
def detalle(contacto_id):
    contacto = buscar_por_id(contacto_id)
    if contacto is None:
        return redirect(url_for("contactos"))
    return render_template("detalle.html", contacto=contacto)


# Editar un contacto existente
@app.route("/editar/<int:contacto_id>", methods=["GET", "POST"])
def editar(contacto_id):
    if "usuario" not in session:
        return redirect(url_for("login"))
    contacto = buscar_por_id(contacto_id)
    if contacto is None:
        return redirect(url_for("contactos"))
    if request.method == "POST":
        nombre = request.form["nombre"]
        apellido = request.form["apellido"]
        telefono = request.form["telefono"]
        correo = request.form["correo"]
        direccion = request.form["direccion"]
        categoria = request.form["categoria"]
        favorito = request.form.get("favorito", "No")
        # Validar datos antes de guardar cambios
        mensaje = validar(nombre, telefono, correo)
        if mensaje != "":
            return render_template("editar.html", contacto=contacto, categorias=CATEGORIAS, error=mensaje)
        libro = load_workbook(ARCHIVO)
        hoja = libro.active
        # Actualizar cada celda con los nuevos datos
        hoja.cell(contacto_id, 1).value = nombre
        hoja.cell(contacto_id, 2).value = apellido
        hoja.cell(contacto_id, 3).value = telefono
        hoja.cell(contacto_id, 4).value = correo
        hoja.cell(contacto_id, 5).value = direccion
        hoja.cell(contacto_id, 6).value = categoria
        hoja.cell(contacto_id, 7).value = favorito
        libro.save(ARCHIVO)
        return redirect(url_for("contactos"))
    return render_template("editar.html", contacto=contacto, categorias=CATEGORIAS)


# Eliminar un contacto
@app.route("/eliminar/<int:contacto_id>", methods=["POST"])
def eliminar(contacto_id):
    libro = load_workbook(ARCHIVO)
    hoja = libro.active
    # Eliminar la fila del contacto
    hoja.delete_rows(contacto_id)
    libro.save(ARCHIVO)
    return redirect(url_for("contactos"))


# Mostrar reportes/estadísticas
@app.route("/reportes")
def reportes():
    lista = leer_contactos()
    total = len(lista)
    # Contar cuántos contactos son favoritos
    favoritos = 0
    for contacto in lista:
        if contacto["favorito"] == "Sí":
            favoritos += 1
    return render_template("reportes.html", total=total, favoritos=favoritos)


# Cerrar sesión
@app.route("/cerrar")
def cerrar():
    session.clear()
    return render_template("cerrar.html")


# Iniciar la aplicación
if __name__ == "__main__":
    crear_excel()
    app.run(debug=True)
